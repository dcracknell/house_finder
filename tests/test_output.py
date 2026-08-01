"""Workbook round-trip and dashboard generation."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from house_finder.output import dashboard, workbook_export, workbook_import
from house_finder.pipeline.dedup import sync_record


def _settings(tmp_path) -> dict:
    return {
        "paths": {
            "xlsx": str(tmp_path / "houses.xlsx"),
            "dashboard_html": str(tmp_path / "dashboard.html"),
        }
    }


def test_export_creates_both_sheets(conn, record_factory, tmp_path):
    sync_record(conn, record_factory(listing_type="sale", fit_score=8.0))
    sync_record(
        conn,
        record_factory(
            property_id="r1", url="http://rent1", listing_type="rent", price=900
        ),
    )
    conn.commit()

    path = workbook_export.export(conn, _settings(tmp_path))
    workbook = load_workbook(path)

    assert "For Sale" in workbook.sheetnames
    assert "To Rent" in workbook.sheetnames
    assert "How to use" in workbook.sheetnames
    assert workbook["For Sale"].max_row == 2
    assert workbook["To Rent"].max_row == 2


def test_export_on_an_empty_database(conn, tmp_path):
    path = workbook_export.export(conn, _settings(tmp_path))
    assert path.exists()
    workbook = load_workbook(path)
    assert workbook["For Sale"].max_row == 1  # headers only


def test_workbook_round_trip_preserves_user_edits(conn, record_factory, tmp_path):
    """The whole point of the workbook: edits must survive a regeneration."""
    settings = _settings(tmp_path)
    record = record_factory()
    sync_record(conn, record)
    conn.commit()

    path = workbook_export.export(conn, settings)

    # Simulate the user editing Status and Notes in Excel.
    workbook = load_workbook(path)
    sheet = workbook["For Sale"]
    headers = [c.value for c in sheet[1]]
    sheet.cell(row=2, column=headers.index("Status") + 1, value="viewing_booked")
    sheet.cell(row=2, column=headers.index("Notes") + 1, value="Viewing Thursday 6pm")
    workbook.save(path)

    counts = workbook_import.import_edits(conn, settings)
    assert counts["status"] == 1
    assert counts["notes"] == 1

    row = conn.execute(
        "SELECT status, notes FROM properties WHERE property_id = ?", (record.property_id,)
    ).fetchone()
    assert row["status"] == "viewing_booked"
    assert row["notes"] == "Viewing Thursday 6pm"


def test_import_does_not_depend_on_hyperlinks(conn, record_factory, tmp_path):
    """Regression: openpyxl cannot read hyperlink targets in read-only mode.

    Matching edits on the Link cell silently imported nothing. Identity must
    come from the hidden Ref column instead.
    """
    settings = _settings(tmp_path)
    record = record_factory()
    sync_record(conn, record)
    conn.commit()
    path = workbook_export.export(conn, settings)

    workbook = load_workbook(path)
    sheet = workbook["For Sale"]
    headers = [c.value for c in sheet[1]]

    assert "Ref" in headers, "the workbook needs a stable identity column"
    ref_cell = sheet.cell(row=2, column=headers.index("Ref") + 1)
    assert ref_cell.value == record.property_id

    # Strip the hyperlink entirely - the import must still work.
    link_cell = sheet.cell(row=2, column=headers.index("Link") + 1)
    link_cell.hyperlink = None
    sheet.cell(row=2, column=headers.index("Status") + 1, value="interested")
    workbook.save(path)

    counts = workbook_import.import_edits(conn, settings)
    assert counts["status"] == 1


def test_ref_column_is_hidden(conn, record_factory, tmp_path):
    sync_record(conn, record_factory())
    conn.commit()
    path = workbook_export.export(conn, _settings(tmp_path))

    sheet = load_workbook(path)["For Sale"]
    headers = [c.value for c in sheet[1]]
    from openpyxl.utils import get_column_letter

    letter = get_column_letter(headers.index("Ref") + 1)
    assert sheet.column_dimensions[letter].hidden is True


def test_import_ignores_an_invalid_status(conn, record_factory, tmp_path):
    settings = _settings(tmp_path)
    sync_record(conn, record_factory())
    conn.commit()
    path = workbook_export.export(conn, settings)

    workbook = load_workbook(path)
    sheet = workbook["For Sale"]
    headers = [c.value for c in sheet[1]]
    sheet.cell(row=2, column=headers.index("Status") + 1, value="not a real status")
    workbook.save(path)

    workbook_import.import_edits(conn, settings)
    row = conn.execute("SELECT status FROM properties").fetchone()
    assert row["status"] == "new"


def test_import_with_no_workbook_is_harmless(conn, tmp_path):
    counts = workbook_import.import_edits(conn, _settings(tmp_path))
    assert counts == {"status": 0, "notes": 0, "skipped": 0}


def test_dashboard_embeds_properties(conn, record_factory, tmp_path):
    sync_record(conn, record_factory(fit_score=8.5, fit_reason="Great match"))
    conn.commit()

    path = dashboard.generate(conn, _settings(tmp_path))
    html = path.read_text(encoding="utf-8")

    assert "leaflet" in html.lower()
    assert "Example Road" in html
    assert "Great match" in html
    assert "const PROPS = [" in html


def test_dashboard_lists_properties_it_cannot_map(conn, record_factory, tmp_path):
    """A property without coordinates must be surfaced, not silently dropped."""
    sync_record(conn, record_factory(lat=None, lon=None))
    conn.commit()

    html = dashboard.generate(conn, _settings(tmp_path)).read_text(encoding="utf-8")
    assert "could not be placed on the map" in html


def test_dashboard_excludes_rejected_and_withdrawn(conn, record_factory, tmp_path):
    sync_record(conn, record_factory(property_id="keep", url="http://keep"))
    sync_record(conn, record_factory(property_id="drop", url="http://drop"))
    conn.execute("UPDATE properties SET status='rejected' WHERE property_id='drop'")
    conn.commit()

    html = dashboard.generate(conn, _settings(tmp_path)).read_text(encoding="utf-8")
    assert "http://keep" in html
    assert "http://drop" not in html


def test_dashboard_property_json_is_valid(conn, record_factory, tmp_path):
    sync_record(conn, record_factory(fit_score=7.0))
    conn.commit()

    html = dashboard.generate(conn, _settings(tmp_path)).read_text(encoding="utf-8")
    start = html.index("const PROPS = ") + len("const PROPS = ")
    end = html.index(";\nconst CENTRE")
    parsed = json.loads(html[start:end])

    assert len(parsed) == 1
    assert parsed[0]["score"] == 7.0
