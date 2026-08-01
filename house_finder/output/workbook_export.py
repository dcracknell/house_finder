"""Regenerate houses.xlsx from SQLite.

The workbook is a VIEW, rebuilt every run - SQLite is the source of truth.
The exception is the Status and Notes columns, which belong to the user and
are imported back before each rebuild (see workbook_import.py).
"""

from __future__ import annotations

import json
import logging
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from house_finder import PROJECT_ROOT
from house_finder.adapters.base import STATUSES

logger = logging.getLogger(__name__)

SHEETS = {"sale": "For Sale", "rent": "To Rent"}

# (header, db column, width). Status and Notes come last so they are easy to
# reach on the right of the sheet.
_COLUMNS = [
    ("Score", "fit_score", 7),
    ("Why", "fit_reason", 46),
    ("Price", "price", 12),
    ("Beds", "bedrooms", 6),
    ("Baths", "bathrooms", 6),
    ("Type", "property_type", 14),
    ("Address", "display_address", 38),
    ("Postcode", "outcode", 10),
    ("Size sqft", "floor_area_sqft", 10),
    ("Tenure", "tenure", 12),
    ("vs Local %", "price_vs_local_pct", 10),
    ("Local sold avg", "local_sold_avg_price", 14),
    ("EPC", "epc_current", 6),
    ("Crime/yr", "crime_incidents_nearby", 9),
    ("Flood", "flood_warnings_nearby", 7),
    ("Broadband", "broadband_max_mbps", 10),
    ("Listed", "first_listed_date", 11),
    ("Agent", "agent_name", 24),
    ("Matched", "matched_criteria", 26),
    ("Link", "url", 16),
    ("Status", "status", 16),
    ("Notes", "notes", 40),
    # Hidden identity column. Edits are matched back to the database on this,
    # NOT on the Link cell: openpyxl cannot read hyperlink targets in
    # read-only mode, so relying on them silently loses every user edit.
    ("Ref", "property_id", 2),
]

_HIDDEN_COLUMNS = {"Ref"}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Score bands -> row tint. Green good, amber middling, grey poor.
_SCORE_FILLS = (
    (8.0, PatternFill("solid", fgColor="C6EFCE")),
    (6.5, PatternFill("solid", fgColor="E2EFDA")),
    (5.0, PatternFill("solid", fgColor="FFF2CC")),
    (0.0, PatternFill("solid", fgColor="F2F2F2")),
)


def xlsx_path(settings: dict) -> Path:
    configured = (settings.get("paths") or {}).get("xlsx", "data/houses.xlsx")
    path = Path(configured)
    return path if path.is_absolute() else PROJECT_ROOT / path


def _score_fill(score) -> PatternFill | None:
    if score is None:
        return None
    for threshold, fill in _SCORE_FILLS:
        if score >= threshold:
            return fill
    return None


def _cell_value(row: sqlite3.Row, column: str):
    value = row[column] if column in row.keys() else None

    if column == "matched_criteria":
        try:
            return ", ".join(json.loads(value or "[]"))
        except (json.JSONDecodeError, TypeError):
            return ""
    if column == "price_vs_local_pct" and value is not None:
        return round(float(value), 1)
    if column == "floor_area_sqft" and value is not None:
        return round(float(value))
    if column == "flood_warnings_nearby":
        return value or 0
    return value


def export(conn: sqlite3.Connection, settings: dict, path: Path | None = None) -> Path:
    """Write the workbook and return where it landed."""
    target = path or xlsx_path(settings)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    status_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUSES) + '"',
        allow_blank=True,
        showDropDown=False,
    )

    total_rows = 0
    for listing_type, sheet_name in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.add_data_validation(status_validation)

        for index, (header, _, width) in enumerate(_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = width
            if header in _HIDDEN_COLUMNS:
                sheet.column_dimensions[letter].hidden = True

        rows = conn.execute(
            """
            SELECT * FROM properties
            WHERE listing_type = ? AND status != 'withdrawn'
            ORDER BY fit_score DESC NULLS LAST, last_seen DESC
            """,
            (listing_type,),
        ).fetchall()

        for row_index, row in enumerate(rows, start=2):
            fill = _score_fill(row["fit_score"])
            for col_index, (_, column, _width) in enumerate(_COLUMNS, start=1):
                cell = sheet.cell(row=row_index, column=col_index)

                if column == "url":
                    cell.value = "View listing"
                    if row["url"]:
                        cell.hyperlink = row["url"]
                        cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = _cell_value(row, column)

                if column in {"fit_reason", "notes", "display_address", "matched_criteria"}:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if column == "price" and row["price"]:
                    cell.number_format = (
                        '"£"#,##0" pcm"' if listing_type == "rent" else '"£"#,##0'
                    )
                if column == "local_sold_avg_price" and row["local_sold_avg_price"]:
                    cell.number_format = '"£"#,##0'
                if column == "price_vs_local_pct" and row["price_vs_local_pct"] is not None:
                    cell.number_format = '+0.0"%";-0.0"%"'
                if fill and column not in {"status", "notes"}:
                    cell.fill = fill

            status_cell = sheet.cell(
                row=row_index,
                column=[c[1] for c in _COLUMNS].index("status") + 1,
            )
            status_validation.add(status_cell)

        sheet.freeze_panes = "A2"
        if rows:
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(_COLUMNS))}{len(rows) + 1}"
            )
        total_rows += len(rows)
        logger.info("export: %s sheet - %d properties", sheet_name, len(rows))

    _write_readme_sheet(workbook)

    workbook.save(target)
    logger.info("export: wrote %d properties to %s", total_rows, target)
    return target


def _write_readme_sheet(workbook: Workbook) -> None:
    """A short how-to-use sheet, so the workbook explains itself."""
    sheet = workbook.create_sheet("How to use")
    sheet.column_dimensions["A"].width = 100
    lines = [
        ("How to use this workbook", True),
        ("", False),
        ("This file is REGENERATED on every run. Two columns are yours and are", False),
        ("never overwritten: Status and Notes. Everything else will be replaced.", False),
        ("", False),
        ("Status - pick from the dropdown:", True),
        ("  new             not looked at yet (set automatically)", False),
        ("  interested      worth a closer look", False),
        ("  viewing_booked  viewing arranged", False),
        ("  viewed          been to see it", False),
        ("  offer_made      offer submitted", False),
        ("  offer_accepted  offer accepted", False),
        ("  rejected        not for you - hidden from future runs for a while", False),
        ("  sold_stc        sold subject to contract (set automatically)", False),
        ("  withdrawn       no longer listed (set automatically)", False),
        ("", False),
        ("Notes - anything you like. Survives every run.", True),
        ("", False),
        ("Score is 0-10 for how well the property matches the criteria in", False),
        ("config/profile.json. 'Why' is the one-line reason for that score.", False),
        ("", False),
        ("vs Local % compares the asking price to what similar properties nearby", True),
        ("actually sold for (HM Land Registry). +15% means 15% above local sold prices.", False),
        ("", False),
        ("Crime/yr is reported incidents within about a mile, per year. A city", False),
        ("centre always scores higher than a village simply because a mile of it", False),
        ("holds more of everything - compare like with like, and treat it as", False),
        ("context rather than a verdict.", False),
        ("", False),
        ("Flood counts flood warnings in force near the property RIGHT NOW.", False),
        ("Zero does not mean the property is not in a flood zone - check properly", False),
        ("before buying.", False),
    ]
    for index, (text, bold) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
