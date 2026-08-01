"""Read the user's Status and Notes edits back into SQLite.

Runs BEFORE the pipeline writes anything, so edits made in Excel since the
last run are never lost when the workbook is regenerated. Excel wins for these
two columns; everything else in the workbook is disposable.
"""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from house_finder.adapters.base import STATUSES
from house_finder.output.workbook_export import SHEETS, xlsx_path

logger = logging.getLogger(__name__)


def _column_indexes(header_row) -> dict[str, int]:
    """Locate the columns by header text, so column order can change safely."""
    wanted = {"ref": "ref", "link": "url", "status": "status", "notes": "notes"}
    found = {}
    for index, cell in enumerate(header_row):
        label = str(cell.value or "").strip().lower()
        if label in wanted:
            found[wanted[label]] = index
    return found


def import_edits(
    conn: sqlite3.Connection, settings: dict, path: Path | None = None
) -> dict[str, int]:
    """Import Status/Notes from the workbook. Returns counts of what changed."""
    target = path or xlsx_path(settings)
    counts = {"status": 0, "notes": 0, "skipped": 0}

    if not target.exists():
        logger.debug("import: no workbook at %s yet", target)
        return counts

    try:
        workbook = load_workbook(target, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - a corrupt workbook must not stop the run
        logger.warning("import: could not read %s: %s", target, exc)
        return counts

    valid_statuses = set(STATUSES)

    try:
        for sheet_name in SHEETS.values():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows()

            try:
                header = next(rows)
            except StopIteration:
                continue

            indexes = _column_indexes(header)
            if "status" not in indexes and "notes" not in indexes:
                continue
            if "ref" not in indexes:
                logger.warning(
                    "import: %s has no Ref column - it predates this version. "
                    "Edits on that sheet cannot be matched; re-run to rebuild it.",
                    sheet_name,
                )
                continue

            ref_index = indexes["ref"]

            for row in rows:
                property_id = None
                if ref_index < len(row):
                    value = row[ref_index].value
                    property_id = str(value).strip() if value else None
                if not property_id:
                    counts["skipped"] += 1
                    continue

                status = None
                notes = None
                if "status" in indexes and indexes["status"] < len(row):
                    raw = row[indexes["status"]].value
                    if raw is not None:
                        candidate = str(raw).strip().lower().replace(" ", "_")
                        if candidate in valid_statuses:
                            status = candidate
                        elif candidate:
                            logger.debug("import: ignoring unknown status %r", raw)
                if "notes" in indexes and indexes["notes"] < len(row):
                    raw = row[indexes["notes"]].value
                    notes = "" if raw is None else str(raw).strip()

                updated = _apply(conn, property_id, status, notes)
                for key in updated:
                    counts[key] += 1
    finally:
        workbook.close()

    conn.commit()
    if counts["status"] or counts["notes"]:
        logger.info(
            "import: applied %d status and %d notes edits from the workbook",
            counts["status"], counts["notes"],
        )
    return counts


def _apply(
    conn: sqlite3.Connection, property_id: str, status: str | None, notes: str | None
) -> list[str]:
    """Write one row's edits, returning which fields actually changed."""
    row = conn.execute(
        "SELECT property_id, status, notes FROM properties WHERE property_id = ?",
        (property_id,),
    ).fetchone()
    if row is None:
        return []

    changed = []
    if status is not None and status != row["status"]:
        conn.execute(
            "UPDATE properties SET status = ? WHERE property_id = ?",
            (status, row["property_id"]),
        )
        changed.append("status")
    if notes is not None and notes != (row["notes"] or ""):
        conn.execute(
            "UPDATE properties SET notes = ? WHERE property_id = ?",
            (notes, row["property_id"]),
        )
        changed.append("notes")
    return changed
